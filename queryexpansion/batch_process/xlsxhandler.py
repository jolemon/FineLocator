from openpyxl import load_workbook

workbook = load_workbook('/Users/lienming/Downloads/defects4j.xlsx')

# booksheet = workbook.active                #获取当前活跃的sheet,默认是第一个sheet

sheets = workbook.sheetnames  # 从名称获取sheet
print(sheets)

projs = []

for sheet in sheets:
    booksheet = workbook[sheet]
    rows = booksheet.rows
    columns = booksheet.columns
    for row in rows:
        if type(row[0].value) is not int:
            continue
        line = [col.value for col in row]

        # print(line)
        proj_id = sheet + '_' + str(line[0])
        proj = {'proj_id': proj_id, 'buggyVersion': line[2], 'mainDir': line[8]}
        print(proj)
        projs.append(proj)

with open('all_buggy_version', 'w') as f:
    f.write('\n'.join(proj['proj_id']+','+proj['buggyVersion']+','+proj['mainDir'] for proj in projs))
